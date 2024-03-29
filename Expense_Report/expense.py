import openpyxl
import xlwt
import datetime


#Create the Month for the 1st column
months = { '01' : 'January' , '02' : 'February' ,  '03' : 'March',
           '04' : 'April',    '05' : 'May',        '06' : 'June',
           '07' : 'July',     '08' : 'August',     '09' : 'September',
           '10' : 'Octomber', '11' : 'November', '12' : 'December'
          }
monthDate = datetime.datetime.now().strftime('%m')
print('The current month is {}'.format(months[monthDate]))



#Load excel file
wb = openpyxl.load_workbook('ExpenseReport2.xlsx')

sheet = wb.get_sheet_by_name('Chirie')

# for cellObj in sheet['A1':'K13']:
#       for cell in cellObj:
#               print(cell.coordinate, str(cell.value))
print(sheet.title)
val = sheet.cell(row=1, column=2).value


#Display values on every row that reside on 1st column
def getTheLastMonth():
    monthRecords = []
    for i in range(1,15):
        if str(sheet.cell(row=i, column=1).value) != 'None':
            print(sheet.cell(row=i, column=1).value)
            monthRecords.append(sheet.cell(row=i, column=1).value)
    return monthRecords[-1]


#Display all the values from sheet('Chirie')
for obj in sheet['A1':'K13']:
    for cell in obj:
        if str(cell.value) != 'None':
            print(cell.coordinate, str(cell.value))

#print(getTheLastMonth())
getTheLastMonth()
